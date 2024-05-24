import pandas as pd
from matplotlib import pyplot as plt
from openpyxl import load_workbook
from reportlab.pdfgen import canvas


def read_dataframe(file_path: str, sheet_name: str) -> pd.DataFrame:
    # Read the named sheet with pandas
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    df.fillna("", inplace=True)
    return df


def generate_pdf(file: str, df: pd.DataFrame) -> str:
    # Create a PDF file using reportlab
    pdf_file_path = f"{file}.pdf"
    c = canvas.Canvas(pdf_file_path, pagesize=(841.89, 595.27))
    width, height = (841.89, 595.27)

    # Create a string representation of the dataframe
    data_str = df.to_string(index=False, max_colwidth=20)

    # Set font and write the data string to the PDF
    c.setFont("Helvetica", 10)
    # c.drawString(30, height - 40, "Excel Sheet Data:")
    text = c.beginText(30, int(height) - 60)
    text.setFont("Helvetica", 8)
    text.textLines(data_str)
    c.drawText(text)
    c.save()
    return pdf_file_path


def generate_jpg(file: str, df: pd.DataFrame, file_path: str) -> str:
    # Load the workbook and sheet with openpyxl for formatting
    wb = load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]
    # Extract cell styles
    styles = {}
    for row in ws.iter_rows():  # type: ignore
        for cell in row:
            if cell.value is not None:
                styles[(cell.row, cell.column)] = cell

    # Generate JPG using matplotlib
    fig, ax = plt.subplots(figsize=(df.shape[1] * 1.5, len(df) // 2))
    ax.axis("tight")
    ax.axis("off")

    table = ax.table(
        cellText=df.values,
        colLabels=df.columns,
        cellLoc="center",
        loc="center",
    )

    # Apply formatting
    for (i, j), cell in table.get_celld().items():
        if (i, j) in styles:
            xl_cell = styles[(i, j)]
            # Set background color if not default
            if (
                xl_cell.fill.start_color.index != "00000000"
                and xl_cell.fill.start_color.rgb is not None
            ):
                color = xl_cell.fill.start_color.rgb[
                    2:
                ]  # Remove the 'FF' prefix
                cell.set_facecolor("#" + color)
            # Set font weight and style
            if xl_cell.font.bold:
                cell.set_text_props(weight="bold")
            if xl_cell.font.italic:
                cell.set_text_props(style="italic")
            # Set font color if specified
            if (
                xl_cell.font.color is not None
                and xl_cell.font.color.rgb is not None
            ):
                color = xl_cell.font.color.rgb[2:]  # Remove the 'FF' prefix
                cell.set_text_props(color="#" + color)

    # Save the JPG file with the same base name as the PDF file
    jpg_file_path = f"{file}.jpg"
    plt.savefig(jpg_file_path, format="jpg", bbox_inches="tight")
    return jpg_file_path


def main(file: str, sheet_name: str) -> None:
    file_path = f"{file}.xlsx"

    df = read_dataframe(file_path, sheet_name)
    pdf_file_path = generate_pdf(file, df)
    jpg_file_path = generate_jpg(file, df, file_path)

    print(f"Generated {pdf_file_path} and {jpg_file_path}")


if __name__ == "__main__":
    # Load the Excel file
    file = "antibiogram"
    sheet_name = "Antibiogram"
    main(file, sheet_name)
